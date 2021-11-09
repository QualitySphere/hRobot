#!/usr/bin/env python3
# -*- coding: utf-8 -*-


# import os
import os

from hrobot.hkeywords import *
from hrobot import hkeywords
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import styles as xl_style
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
import robot
from robot import libraries
from robot.api import logger
# from selenium import webdriver
# import requests
# import platform
# import paramiko
import inspect
import json
import yaml
# import time
# import re


def print_info(info_string):
    logger.info(info_string, html=True, also_console=True)
    return True


class HRobot(object):
    def __init__(self):
        self.work_dir = os.path.abspath('.')
        self.robot_dir = os.path.basename(self.work_dir)
        self.project_file = '.hrobot'
        self.hrobot_keywords_robot_file = 'hrobot.robot'
        self.output_dir = 'output'
        self.testcases_dir = 'testcases'
        self.testcases_file = 'suite.xlsx'
        self.testcases_sheets = [u"用例", u"变量", u"前置", u"后置", u"可用关键字"]
        self.testcases_headers = {
            u"用例": [u"用例标题", u"标签", u"用例描述", u"关键字库", u"关键字", u"参数"],
            u"变量": [u"变量类型", u"变量名", u"变量值"],
            u"前置": [u"关键字库", u"关键字", u"参数"],
            u"后置": [u"关键字库", u"关键字", u"参数"],
            u"可用关键字": [u"关键字库", u"关键字", u"参数"],
        }
        self.testcases_col_width = {
            u"用例": [24, 14, 32, 14, 24, 24, 24, 24, 24, 24],
            u"变量": [12, 32, 32],
            u"前置": [14, 24, 24, 24, 24, 24, 24],
            u"后置": [14, 24, 24, 24, 24, 24, 24],
            u"可用关键字": [14, 24, 24, 24, 24, 24, 24],
        }
        self.keywords_dir = "keywords"
        self.keywords_file = "keywords.xlsx"
        self.keywords_sheets = [u"自定义关键字"]
        self.keywords_headers = {
            u"自定义关键字": [u"自定义关键字", u"入参", u"返回", u"关键字库", u"关键字", u"参数"],
        }
        self.keywords_col_width = {
            u"自定义关键字": [24, 14, 32, 14, 24, 24, 24, 24, 24, 24],
        }
        self.variables_dir = "variables"
        self.variables_file = "variables.xlsx"
        self.variables_sheets = [u"变量"]
        self.variables_headers = {
            u"变量": [u"变量类型", u"变量名", u"变量值"],
        }
        self.variables_col_width = {
            u"变量": [12, 32, 32],
        }
        self.book_header_font = xl_style.Font(name=u'黑体', size=12, bold=True)
        self.book_header_pattern_fill = xl_style.PatternFill(patternType='solid', fgColor='bfbfbf')
        self.book_header_protection = xl_style.Protection()
        self.book_font = xl_style.Font(name=u'黑体', size=12)
        self.book_alignment = xl_style.Alignment(horizontal='left', vertical='center')

    @staticmethod
    def __reload_hrobot_keywords_to_xl_sheet(xl_sheet):
        # 提取出 hRobot 中可用的关键字列表
        keyword_index = list()
        keyword_row = 0
        # keyword_start_col = 0
        # keyword_end_col = 0
        for kw_cls_name in hkeywords.__dict__.keys():
            if hkeywords.__dict__[kw_cls_name].__doc__ != u"关键字":
                continue
            # print(kw_cls_name)
            keyword_start_col = xl_sheet.max_row + 1
            kw_cls = hkeywords.__dict__[kw_cls_name]()
            keyword_lib = inspect.getdoc(kw_cls.__init__)
            for kw_fun_name in kw_cls.__dir__():
                if kw_fun_name.startswith('_'):
                    continue
                keyword_obj = kw_cls.__getattribute__(kw_fun_name)
                keyword_name = inspect.getdoc(keyword_obj)
                keyword_row += 1
                keyword_args = list()
                fun_args = inspect.getfullargspec(keyword_obj)
                for fun_arg in fun_args.args[1:]:
                    keyword_args.append('${%s}' % fun_arg)
                arg_defaults = fun_args.defaults
                if arg_defaults:
                    for i in range(-1, -len(arg_defaults) - 1, -1):
                        keyword_args[i] = '%s=%s' % (keyword_args[i], str(arg_defaults[i]))
                xl_sheet.append([keyword_lib, keyword_name] + keyword_args)
                print(u'加载可用关键字 %s.%s' % (keyword_lib, keyword_name))
            keyword_end_col = xl_sheet.max_row
            keyword_index.append([keyword_lib, keyword_start_col, keyword_end_col])
        return keyword_index

    @staticmethod
    def __define_names_for_keywords(xl_book, xl_index):
        # index_end_col = 2
        for index_name, index_start_col, index_end_col in xl_index:
            defined_name = DefinedName(
                index_name,
                attr_text=u"可用关键字!$B$%s:$B$%s" % (index_start_col, index_end_col)
            )
            if xl_book.defined_names.get(index_name):
                xl_book.defined_names.delete(index_name)
            xl_book.defined_names.append(defined_name)
        # index_name = u'可用关键字'
        # defined_name = DefinedName(
        #     index_name,
        #     attr_text=u"可用关键字!$B$2:$B$%s" % index_end_col
        # )
        # if xl_book.defined_names.get(index_name):
        #     xl_book.defined_names.delete(index_name)
        # xl_book.defined_names.append(defined_name)

    @staticmethod
    def __set_row_height(xl_sheet, max_row):
        """
        设置 Sheet 的行高
        :param xl_sheet:
        :param max_row:
        :return:
        """
        for i in range(1, max_row):
            xl_sheet.row_dimensions[i].height = 26
        return True

    @staticmethod
    def __set_col_width(xl_sheet, width_list):
        """
        设置 Sheet 的列宽
        :param xl_sheet:
        :param width_list:
        :return:
        """
        _col = 65
        for col_width in width_list:
            xl_sheet.column_dimensions[chr(_col)].width = col_width
            _col += 1

    def __set_sheet_header(self, xl_sheet):
        """
        设置 Sheet 的表头格式
        :param xl_sheet:
        :return:
        """
        for _cell in xl_sheet[1]:
            _cell.font = self.book_header_font
            _cell.fill = self.book_header_pattern_fill
        return True

    def __set_sheet_cell(self, xl_sheet, cols):
        """
        设置 Sheet 单元格格式
        :param xl_sheet:
        :param cols:
        :return:
        """
        for _col in range(65, 65 + len(cols)):
            for _cell in xl_sheet[chr(_col)]:
                _cell.font = self.book_font
                _cell.alignment = self.book_alignment

    @staticmethod
    def __set_sheet_data_validation(xl_sheet, xl_lib_col, xl_kw_col, kw_libs):
        """
        设置 Sheet 关键字库和关键字的数据验证
        :param xl_sheet:
        :param xl_lib_col:
        :param xl_kw_col:
        :param kw_libs:
        :return:
        """
        for _validation in xl_sheet.data_validations.dataValidation:
            xl_sheet.data_validations.dataValidation.remove(_validation)
        max_row = xl_sheet.max_row if xl_sheet.max_row >= 500 else 500
        validation_list = set()
        for item in kw_libs:
            validation_list.add(item[0])
        lib_data_validation = DataValidation(
            type='list',
            formula1='"%s"' % ','.join(validation_list),
            allow_blank=True
        )
        lib_data_validation.add('%s2:%s%s' % (xl_lib_col, xl_lib_col, max_row))
        xl_sheet.add_data_validation(lib_data_validation)
        for xl_kw_row in range(2, max_row):
            kw_data_validation = DataValidation(
                type='list',
                formula1='INDIRECT(%s%s)' % (xl_lib_col, xl_kw_row),
                allow_blank=True
            )
            kw_data_validation.add('%s%s' % (xl_kw_col, xl_kw_row))
            xl_sheet.add_data_validation(kw_data_validation)
        return True

    @staticmethod
    def __set_sheet_data_validation_for_variable_type(xl_sheet):
        var_type_data_validation = DataValidation(
            type='list',
            formula1='"str,int,list,dict"',
            allow_blank=True
        )
        var_type_data_validation.add('A2:A50')
        xl_sheet.add_data_validation(var_type_data_validation)

    def generate_testcase_xl(self, xl_file):
        book = Workbook()
        # 开始 定义 Sheet 用例
        sheet_name = self.testcases_sheets[0]
        sheet_case = book.create_sheet(sheet_name, 0)
        self.__set_row_height(sheet_case, 500)
        self.__set_col_width(sheet_case, self.testcases_col_width[sheet_name])
        sheet_case.append(self.testcases_headers[sheet_name])
        # Demo Start #
        sheet_case.append([u'Demo演示', 'demo', u'用于给初学者的展示', u'内置', u'打印日志', u'这是一个演示用的用例'])
        sheet_case.append(['', '', '', u'接口', 'GET', 'https://QualitySphere.gitee.io'])
        sheet_case.append(['', '', '', u'接口', u'响应.断言', 'status_code', '=', '200'])
        # Demo End #
        # 开始设置单元格样式
        self.__set_sheet_cell(sheet_case, self.testcases_col_width[sheet_name])
        self.__set_sheet_header(sheet_case)
        sheet_case.freeze_panes = 'F2'
        # 完成设置单元格样式
        # 完成 定义 Sheet 用例

        # 开始 定义 Sheet 变量
        sheet_name = self.testcases_sheets[1]
        sheet_variables = book.create_sheet(sheet_name, 1)
        self.__set_row_height(sheet_variables, 50)
        self.__set_col_width(sheet_variables, self.testcases_col_width[sheet_name])
        sheet_variables.append(self.testcases_headers[sheet_name])
        # 开始设置单元格样式
        self.__set_sheet_cell(sheet_variables, self.testcases_col_width[sheet_name])
        self.__set_sheet_header(sheet_variables)
        # 完成设置单元格样式
        # 完成 定义 Sheet 变量

        # 开始 定义 Sheet 前置
        sheet_name = self.testcases_sheets[2]
        sheet_setup = book.create_sheet(sheet_name, 2)
        self.__set_row_height(sheet_setup, 50)
        self.__set_col_width(sheet_setup, self.testcases_col_width[sheet_name])
        sheet_setup.append(self.testcases_headers[sheet_name])
        sheet_setup.append([u'内置', u'打印日志', u'测试用例集执行前的准备工作'])
        # 开始设置单元格样式
        self.__set_sheet_cell(sheet_setup, self.testcases_col_width[sheet_name])
        self.__set_sheet_header(sheet_setup)
        # 完成设置单元格样式
        # 完成 定义 Sheet 前置

        # 开始 定义 Sheet 后置
        sheet_name = self.testcases_sheets[3]
        sheet_teardown = book.create_sheet(sheet_name, 3)
        self.__set_row_height(sheet_teardown, 50)
        self.__set_col_width(sheet_teardown, self.testcases_col_width[sheet_name])
        sheet_teardown.append(self.testcases_headers[sheet_name])
        sheet_teardown.append([u'内置', u'打印日志', u'测试用例集执行前的清理工作'])
        # 开始设置单元格样式
        self.__set_sheet_cell(sheet_teardown, self.testcases_col_width[sheet_name])
        self.__set_sheet_header(sheet_teardown)
        # 完成设置单元格样式
        # 完成 定义 Sheet 后置

        # 开始 定义 Sheet 可用关键字
        sheet_name = self.testcases_sheets[4]
        sheet_keyword = book.create_sheet(sheet_name, 4)
        self.__set_col_width(sheet_keyword, self.testcases_col_width[sheet_name])
        sheet_keyword.append(self.testcases_headers[sheet_name])
        # 提取出 hRobot 中可用的关键字列表
        keyword_index = self.__reload_hrobot_keywords_to_xl_sheet(sheet_keyword)
        # 提取完成
        # 开始设置单元格样式
        self.__set_sheet_cell(sheet_keyword, self.testcases_col_width[sheet_name])
        self.__set_sheet_header(sheet_keyword)
        self.__set_row_height(sheet_keyword, sheet_keyword.max_row + 1)
        # 完成设置单元格样式
        # 完成 定义 Sheet 可用关键字

        # 添加数据验证配置和定义名称
        self.__define_names_for_keywords(book, keyword_index)
        self.__set_sheet_data_validation(sheet_case, 'D', 'E', keyword_index)
        self.__set_sheet_data_validation(sheet_setup, 'A', 'B', keyword_index)
        self.__set_sheet_data_validation(sheet_teardown, 'A', 'B', keyword_index)
        self.__set_sheet_data_validation_for_variable_type(sheet_variables)
        # 完成数据验证配置和定义名称
        book.save(xl_file)
        book.close()

    def generate_variable_xl(self, xl_file):
        book = Workbook()
        sheet_name = self.variables_sheets[0]
        sheet = book.create_sheet(sheet_name, 0)
        sheet.append(self.variables_headers[sheet_name])
        self.__set_row_height(sheet, 50)
        self.__set_col_width(sheet, self.variables_col_width[sheet_name])
        # 开始设置单元格样式
        self.__set_sheet_cell(sheet, self.variables_col_width[sheet_name])
        self.__set_sheet_header(sheet)
        # 完成设置单元格样式
        self.__set_sheet_data_validation_for_variable_type(sheet)
        book.save(xl_file)
        book.close()

    def generate_keyword_xl(self, xl_file):
        book = Workbook()
        sheet_name = self.keywords_sheets[0]
        sheet = book.create_sheet(sheet_name, 0)
        sheet.append(self.keywords_headers[sheet_name])
        book.save(xl_file)
        book.close()

    def init_project(self, cmd_args: dict):
        """
        初始化项目目录，若存在则终止
        :param: cmd_args
        :return:
        """
        project_path = os.path.join(self.work_dir, cmd_args['project'])
        if os.path.exists(project_path):
            print(u"项目目录 %s 已经存在" % project_path)
            exit(1)
        os.mkdir(project_path)
        os.mkdir(os.path.join(project_path, self.testcases_dir))
        os.mkdir(os.path.join(project_path, self.variables_dir))
        os.mkdir(os.path.join(project_path, self.keywords_dir))
        self.generate_testcase_xl(os.path.join(project_path, self.testcases_dir, self.testcases_file))
        self.generate_variable_xl(os.path.join(project_path, self.variables_dir, self.variables_file))
        self.generate_keyword_xl(os.path.join(project_path, self.keywords_dir, self.keywords_file))
        with open(os.path.join(project_path, self.project_file), 'w', encoding='utf-8') as f:
            f.write(yaml.safe_dump({
                "PROJECT": cmd_args['project'],
            }))
        with open(os.path.join(project_path, '.gitignore'), 'w', encoding='utf-8') as f:
            f.write('\n'.join([
                '.DS_Store',
                '__pycache__/',
                '.pytest_cache__/',
                '.idea/',
                'robotframework/'
            ]))
        return True

    # @staticmethod
    # def __smart_content(content, robot_variables):
    #     new_content = content
    #     var_pattern = re.compile("{{[a-zA-Z0-9 _-]*}}")
    #     if robot_variables:
    #         for var_key_str in var_pattern.findall(content):
    #             var_key = var_key_str.strip('{{').strip('}}').strip()
    #             try:
    #                 # 若变量能在参数中找到,就处理成 robot 变量格式
    #                 if robot_variables[var_key]['type'] == 'list':
    #                     new_content = content.replace(var_key_str, '@{%s}' % var_key)
    #                 elif robot_variables[var_key]['type'] == 'dict':
    #                     new_content = content.replace(var_key_str, '&{%s}' % var_key)
    #                 else:
    #                     new_content = content.replace(var_key_str, '${%s}' % var_key)
    #             except KeyError:
    #                 # 如果找不到就不处理
    #                 print(u'未在参数中找到该变量值')
    #     return new_content

    @staticmethod
    def __smart_keyword_and_arguments(kw_name, kw_args):
        key_value_keywords = {
            u"设置变量": 'set test variable',
            u"设置用例集变量": 'set suite variable',
            u"设置全局变量": 'set global variable',
            u"当前时间": 'set global variable',
        }
        no_arg_keywords = {
            u"当前时间戳": 'hrobot get current timestamp',
        }
        advanced_keywords = {
            u"响应.取值": u'响应.取值',
        }
        if kw_name in key_value_keywords.keys():
            new_kw_name = key_value_keywords[kw_name]
            new_kw_args = list()
            new_kw_args.append('${%s}' % kw_args[0])
            new_kw_args.append(kw_args[1])
        elif kw_name in no_arg_keywords.keys():
            new_kw_name = '${%s}=' % kw_args[-1]
            new_kw_args = list()
            new_kw_args.append(no_arg_keywords[kw_name])
        elif kw_name in advanced_keywords.keys():
            new_kw_name = '${%s}=' % kw_args[-1]
            new_kw_args = list()
            new_kw_args.append(advanced_keywords[kw_name])
            for kw_arg in kw_args:
                new_kw_args.append(kw_arg)
        else:
            new_kw_name = kw_name
            new_kw_args = kw_args
        return new_kw_name, new_kw_args

    def xl_to_robot_case(self, xl_file):
        """
        excel 文件转换为 RobotFramework 用例文件 .robot
        :param xl_file:
        :return:
        """
        book = load_workbook(xl_file)
        robot_file_name_prefix = os.path.basename(xl_file).split('.')[0]
        robot_file = os.path.join(
            self.work_dir,
            self.robot_dir,
            self.testcases_dir,
            '%s' % robot_file_name_prefix
        )
        robot_json = {
            'settings': {
                'documentation': robot_file,
                'resource': set(),
                'suite_setup': u'用例集前置',
                'suite_teardown': u'用例集后置',
                'test_setup': set(),
                'test_teardown': set(),
            },
            'variables': {},
            'testcases': [],
            'keywords': []
        }
        # hrobot_keywords_file = os.path.join('..', self.env['KEYWORDS_DIR'], self.env['HROBOT_KEYWORDS_ROBOT_FILE'])
        # robot_json['settings']['resource'].add(hrobot_keywords_file)

        # 开始解析 sheet 变量
        sheet_variables = book[u"变量"]
        sheet_header = dict()
        col_num = 0
        for col_name in sheet_variables[1]:
            sheet_header[col_name.value] = col_num
            col_num += 1
        for row_data in list(sheet_variables.rows)[1:]:
            var_key = row_data[sheet_header[u'变量名']].value
            var_type = row_data[sheet_header[u'变量类型']].value
            var_value = row_data[sheet_header[u'变量值']].value
            if var_type in ['str']:
                var_type = 'str'
                var_value = str(var_value)
            elif var_type in ['int']:
                var_type = 'int'
                var_value = int(float(var_value))
            elif var_type in ['list']:
                var_type = 'list'
                var_value = json.dumps(var_value)
            elif var_type in ['dict']:
                var_type = 'dict'
                var_value = json.dumps(var_value)
            robot_json['variables'][var_key] = {
                "type": var_type,
                "value": var_value
            }
            # logger.info(robot_json, also_console=True)
        # 解析 sheet 变量 完成

        # 开始解析 sheet 用例
        sheet_case = book[u'用例']
        # <开始表头解析> 第0行是表头，处理成字典格式，表头与列号的对应关系，好在后续用例解析的时候灵活使用
        sheet_header = dict()
        col_num = 0
        for col_name in sheet_case[1]:
            sheet_header[col_name.value] = col_num
            col_num += 1
        # <完成表头解析>
        # 第1行开始是测试用例数据
        for row_data in list(sheet_case.rows)[1:]:
            # 开始处理 用例标题 和 用例描述 : Excel 表头是"用例标题"和"用例描述"的列号单元格中数据
            case_title = row_data[sheet_header[u'用例标题']].value
            case_description = row_data[sheet_header[u'用例描述']].value
            if not case_description:
                case_description = ''
            case_tags = row_data[sheet_header[u'标签']].value
            if not case_tags:
                case_tags = ''
            # 如果测试用例数据尚无记录，或者A列不为空且用例标题与记录中最后一个不同，就初始化一个新的用例记录，虽然可以简单粗暴的在 .robot 加空行，但似乎这样处理更美观，待后续再看看有无更好的方案
            if len(robot_json['testcases']) == 0 or \
                    case_title and case_title != robot_json['testcases'][-1]['title']:
                print(u'发现测试用例 %s' % case_title)
                robot_json['testcases'].append({
                    'title': case_title,
                    'description': case_description,
                    'tags': case_tags.split(','),
                    'steps': []
                })
            # 完成处理 用例标题 和 用例描述
        # Excel 中表头是"关键字库" 和 表头是"关键字" 的单元格数据拼装出真正的关键字
            step_kw_lib = row_data[sheet_header[u'关键字库']].value
            step_kw_name = row_data[sheet_header[u'关键字']].value
            # step_kw = '.'.join([step_kw_lib, step_kw_name])
            if step_kw_name:
                robot_json['settings']['resource'].add(os.path.join('..', 'keywords', '%s.robot' % step_kw_lib))
                # print(robot_json['settings']['resource'])
                # Excel 中表头从"参数"开始后面都是参数，添加到用例记录的最后一个用例中去
                step_args = list()
                for step_arg in row_data[sheet_header[u'参数']:]:
                    # logger.info(step_arg.value, also_console=True)
                    # logger.info(robot_json['variables'], also_console=True)
                    # step_args.append('%s' % self.__smart_content(str(step_arg.value), robot_json['variables']))
                    if step_arg.value:
                        step_args.append('%s' % step_arg.value)
                step_kw_name, step_args = self.__smart_keyword_and_arguments(step_kw_name, step_args)
                robot_json['testcases'][-1]['steps'].append('\t'.join([
                    step_kw_name,
                    '\t'.join(step_args)
                ]))
        # 解析 sheet 用例 完成

        # 开始解析 sheet 前置
        sheet_setup = book[u'前置']
        # sheet_setup = book.sheet_by_name(u'前置')
        sheet_header = dict()
        col_num = 0
        for col_name in sheet_setup[1]:
            sheet_header[col_name.value] = col_num
            col_num += 1
        suite_setup_steps = list()
        for row_data in list(sheet_setup.rows)[1:]:
            step_kw_lib = row_data[sheet_header[u'关键字库']].value
            step_kw_name = row_data[sheet_header[u'关键字']].value
            # step_kw = '.'.join([step_kw_type, step_kw_name])
            if step_kw_name:
                robot_json['settings']['resource'].add(os.path.join('..', 'keywords', '%s.robot' % step_kw_lib))
                step_args = list()
                for step_arg in row_data[sheet_header[u'参数']:]:
                    if step_arg.value:
                        step_args.append(str(step_arg.value))
                step_kw_name, step_args = self.__smart_keyword_and_arguments(step_kw_name, step_args)
                suite_setup_steps.append('\t'.join([step_kw_name, '\t'.join(step_args)]))
        robot_json['keywords'].append({
            robot_json['settings']['suite_setup']: '\t' + '\n\t'.join(suite_setup_steps)
        })
        # 解析 sheet 前置 完成

        # 开始解析 sheet 后置
        sheet_teardown = book[u'后置']
        # sheet_teardown = book.sheet_by_name(u'后置')
        sheet_header = dict()
        col_num = 0
        for col_name in sheet_teardown[1]:
            sheet_header[col_name.value] = col_num
            col_num += 1
        suite_teardown_steps = list()
        for row_data in list(sheet_teardown.rows)[1:]:
            step_kw_lib = row_data[sheet_header[u'关键字库']].value
            step_kw_name = row_data[sheet_header[u'关键字']].value
            # step_kw = '.'.join([step_kw_type, step_kw_name])
            if step_kw_name:
                robot_json['settings']['resource'].add(os.path.join('..', 'keywords', '%s.robot' % step_kw_lib))
                step_args = list()
                for step_arg in row_data[sheet_header[u'参数']:]:
                    if step_arg.value:
                        step_args.append(str(step_arg.value))
                step_kw_name, step_args = self.__smart_keyword_and_arguments(step_kw_name, step_args)
                suite_teardown_steps.append('\t'.join([step_kw_name, '\t'.join(step_args)]))
        robot_json['keywords'].append({
            robot_json['settings']['suite_teardown']: '\t' + '\n\t'.join(suite_teardown_steps)
        })
        # 解析 sheet 后置 完成

        robot_libs = '\nResource         '.join(robot_json['settings']['resource'])
        robot_settings = '\n'.join([
            '*** Settings ***',
            u'Documentation    %s' % robot_json['settings']['documentation'],
            u'Resource         %s' % robot_libs,
            u'Suite Setup      %s' % robot_json['settings']['suite_setup'],
            u'Suite Teardown   %s' % robot_json['settings']['suite_teardown'],
        ])
        robot_variables = '*** Variables ***'
        for var_item_key, var_item_value in robot_json['variables'].items():
            var_item_value_type = var_item_value['type']
            var_item_value_str = var_item_value['value']
            print_info(u'加载变量 %s \t: %s' % (var_item_key, var_item_value_str))
            if var_item_value_type in ['str', 'int']:
                var_item_string = '${%s}\t%s' % (var_item_key, var_item_value_str)
            elif var_item_value_type == 'list':
                var_item_string = '@{%s}\t%s' % (var_item_key, var_item_value_str)
            elif var_item_value_type == 'dict':
                var_item_string = '&{%s}\t%s' % (var_item_key, var_item_value_str)
            else:
                var_item_string = '${%s}\t%s' % (var_item_key, var_item_value_str)
            robot_variables = '\n'.join([
                robot_variables,
                var_item_string
            ])
            # logger.info('%s' % robot_variables, also_console=True)

        robot_testcases = '*** Test Cases ***'
        for tc in robot_json['testcases']:
            robot_steps = '\n\t'.join(tc['steps'])
            robot_testcases = '\n'.join([
                robot_testcases,
                tc['title'],
                u'\t[Documentation]\t%s' % tc['description'],
                u'\t[Tags]\t%s' % '\t'.join(tc['tags']),
                u'\t%s' % robot_steps
            ])
        robot_keywords = '*** Keywords ***'
        for kw in robot_json['keywords']:
            for kw_key, kw_value in kw.items():
                robot_keywords = '\n'.join([
                    robot_keywords,
                    kw_key,
                    kw_value
                ])
        robot_content = '\n'.join([
            robot_settings,
            robot_variables,
            robot_testcases,
            robot_keywords
        ])
        with open('%s.robot' % robot_file, 'w', encoding='utf-8') as f:
            f.write(robot_content)

    @staticmethod
    def __cls_to_robot_keywords(kw_lib):
        """
        把 Class 转换成 RobotFramework 关键字
        :return:
        """
        rbt_kws = list()
        # pprint('%s' % kw_lib.__dir__())
        for fun_name in kw_lib.__dir__():
            if fun_name.startswith('_'):
                continue
            fun_obj = kw_lib.__getattribute__(fun_name)
            if not fun_obj:
                continue
            kw_name = inspect.getdoc(fun_obj)
            fun_args = inspect.getfullargspec(fun_obj)
            kw_args = list()
            for fun_arg in fun_args.args[1:]:
                kw_args.append(['${%s}' % fun_arg])
            arg_defaults = fun_args.defaults
            if arg_defaults:
                for i in range(-1, -len(arg_defaults) - 1, -1):
                    kw_args[i].append(str(arg_defaults[i]))
            rf_kw_args = list()
            rf_fun_args = list()
            for kv in kw_args:
                rf_kw_args.append('='.join(kv))
                rf_fun_args.append(kv[0])
            # pprint('Keywords args %s' % rf_kw_args)
            # pprint('Function name %s' % fun_name)
            # pprint('Function args %s' % rf_fun_args)
            rbt_kws.append('\n'.join([
                kw_name,
                '    [Arguments]    %s' % '    '.join(rf_kw_args),
                '    [Return]       ${KEYWORD_RETURN}',
                '    ${KEYWORD_RETURN}    %s    %s' % (fun_name, '    '.join(rf_fun_args))
            ]))
            # pprint('RF Keywords %s' % rbt_kws)
            # print('\n')
        return rbt_kws

    def cls_to_robot_builtin_keyword(self):
        """
        把 RobotFramework 的关键字转换成中文到 .robot 文件
        :return:
        """
        for kw_cls_name in hkeywords.__dict__.keys():
            if hkeywords.__dict__[kw_cls_name].__doc__ != u"关键字":
                continue
            kw_cls = hkeywords.__dict__[kw_cls_name]()
            keyword_lib = inspect.getdoc(kw_cls.__init__)
            # print(u'开始处理 %s' % keyword_lib)
            # for kw_fun_name in kw_cls.__dir__():
            #     if kw_fun_name.startswith('_'):
            #         continue
            #     keyword_name = inspect.getdoc(kw_cls.__getattribute__(kw_fun_name))
            #     print(u'发现可用关键字 %s.%s' % (keyword_lib, keyword_name))
            robot_file = os.path.join(
                self.work_dir,
                self.robot_dir,
                self.keywords_dir,
                "%s.robot" % keyword_lib
            )
            robot_keywords = self.__cls_to_robot_keywords(kw_cls)
            if kw_cls_name in list(libraries.STDLIBS):
                robot_keywords_lib = kw_cls_name
            else:
                robot_keywords_lib = 'hrobot.hkeywords.%s' % kw_cls_name
            robot_content = '\n'.join([
                u'*** Settings ***',
                u'Documentation    Hybrid Robot Keywords',
                u'Library          %s' % robot_keywords_lib,
                u'\n',
                u'*** Keywords ***',
                u'\n'.join(robot_keywords),
                u'\n',
            ])
            with open(robot_file, 'w', encoding='utf-8') as f:
                f.write(robot_content)

    def xl_to_robot_keyword(self, xl_file):
        pass

    def xl_to_robot_variable(self, xl_file):
        robot_file_name_prefix = os.path.basename(xl_file).split('.')[0]
        robot_file = os.path.join(
            self.work_dir,
            self.robot_dir,
            self.variables_dir,
            '%s' % robot_file_name_prefix
        )
        robot_json = dict()
        book = load_workbook(xl_file)
        # 开始解析 sheet 变量
        sheet_variables = book[u"变量"]
        sheet_header = dict()
        col_num = 0
        for col_name in sheet_variables[1]:
            sheet_header[col_name.value] = col_num
            col_num += 1
        for row_data in list(sheet_variables.rows)[1:]:
            var_key = row_data[sheet_header[u'变量名']].value
            var_type = row_data[sheet_header[u'变量类型']].value
            var_value = row_data[sheet_header[u'变量值']].value
            if var_type in ['str']:
                var_value = str(var_value)
            elif var_type in ['int']:
                var_value = int(float(var_value))
            elif var_type in ['list']:
                var_value = json.dumps(var_value)
            elif var_type in ['dict']:
                var_value = json.dumps(var_value)
            print_info(u'加载变量 %s : %s' % (var_key, var_value))
            robot_json[var_key] = var_value
        # 解析 sheet 变量 完成
        with open('%s.yaml' % robot_file, 'w', encoding='utf-8') as f:
            f.write(yaml.safe_dump(robot_json))
        book.close()

    def run_project(self, cmd_args: dict):
        """
        执行项目测试，先转换成 RobotFramework 结构目录文件，然后调用 robot 执行
        :param: cmd_args
        :return:
        """
        if not os.path.exists(self.project_file):
            print(u'这不是一个 hRobot 项目目录')
            return False
        robot_path = os.path.join(self.work_dir, self.robot_dir)
        os.system('rm -rf %s' % robot_path)
        os.mkdir(robot_path)
        os.mkdir(os.path.join(robot_path, self.testcases_dir))
        os.mkdir(os.path.join(robot_path, self.keywords_dir))
        os.mkdir(os.path.join(robot_path, self.variables_dir))
        self.cls_to_robot_builtin_keyword()
        for case_file in os.listdir(os.path.join(self.work_dir, self.testcases_dir)):
            if str(case_file).split('.')[-1] not in ['xlsx', 'xlsm', 'xltx', 'xltm'] or str(case_file).startswith('~'):
                continue
            print(u'开始解析 %s' % case_file)
            xl_case_file = os.path.join(self.work_dir, self.testcases_dir, case_file)
            self.xl_to_robot_case(xl_case_file)
            # 删除重建 Excel 中 可用关键字 sheet ，重新加载 可用关键字
            book = load_workbook(xl_case_file)
            sheet_keyword = book[u'可用关键字']
            book.remove(sheet_keyword)
            sheet_keyword = book.create_sheet(u'可用关键字', 4)
            sheet_keyword.append(self.testcases_headers[u'可用关键字'])
            keyword_index = self.__reload_hrobot_keywords_to_xl_sheet(sheet_keyword)
            # 添加数据验证定义名称
            self.__define_names_for_keywords(book, keyword_index)
            # 添加数据验证配置
            self.__set_sheet_data_validation(book[u'用例'], 'D', 'E', keyword_index)
            self.__set_sheet_data_validation(book[u'前置'], 'A', 'B', keyword_index)
            self.__set_sheet_data_validation(book[u'后置'], 'A', 'B', keyword_index)
            self.__set_sheet_data_validation_for_variable_type(book[u'变量'])
            # 完成数据验证配置和定义名称
            self.__set_sheet_cell(sheet_keyword, self.testcases_col_width[u'可用关键字'])
            self.__set_sheet_header(sheet_keyword)
            self.__set_row_height(sheet_keyword, sheet_keyword.max_row + 1)
            self.__set_col_width(sheet_keyword, self.testcases_col_width[u'可用关键字'])
            self.__set_sheet_cell(book[u'用例'], self.testcases_col_width[u'用例'])
            self.__set_sheet_header(book[u'用例'])
            self.__set_row_height(book[u'用例'], book[u'用例'].max_row + 1)
            self.__set_col_width(book[u'用例'], self.testcases_col_width[u'用例'])
            self.__set_sheet_cell(book[u'变量'], self.testcases_col_width[u'变量'])
            self.__set_sheet_header(book[u'变量'])
            self.__set_row_height(book[u'变量'], book[u'变量'].max_row + 1)
            self.__set_col_width(book[u'变量'], self.testcases_col_width[u'变量'])
            self.__set_sheet_cell(book[u'前置'], self.testcases_col_width[u'前置'])
            self.__set_sheet_header(book[u'前置'])
            self.__set_row_height(book[u'前置'], book[u'前置'].max_row + 1)
            self.__set_col_width(book[u'前置'], self.testcases_col_width[u'前置'])
            self.__set_sheet_cell(book[u'后置'], self.testcases_col_width[u'后置'])
            self.__set_sheet_header(book[u'后置'])
            self.__set_row_height(book[u'后置'], book[u'后置'].max_row + 1)
            self.__set_col_width(book[u'后置'], self.testcases_col_width[u'后置'])
            book.save(xl_case_file)
            book.close()
            # 更新完成
        allure_results_dir = os.path.join(robot_path, self.output_dir, 'allure-results')
        # 开始解析环境变量
        try:
            xl_files = os.listdir(os.path.join(self.work_dir, self.variables_dir))
        except FileNotFoundError:
            xl_files = list()
        for variable_file in xl_files:
            if str(variable_file).split('.')[-1] not in ['xlsx', 'xlsm', 'xltx', 'xltm'] or str(variable_file).startswith('~'):
                continue
            print(u'开始解析 %s' % variable_file)
            xl_variable_file = os.path.join(self.work_dir, self.variables_dir, variable_file)
            self.xl_to_robot_variable(xl_variable_file)
        variable_files = list()
        try:
            yaml_files = os.listdir(os.path.join(self.robot_dir, self.variables_dir))
        except FileNotFoundError:
            yaml_files = list()
        for _file in yaml_files:
            variable_files.append(os.path.join(self.robot_dir, self.variables_dir, _file))
        # 完成解析环境变量
        if cmd_args['suite'] and cmd_args['case']:
            robot.run(
                self.robot_dir,
                consolewidth=80,
                consolecolors='on',
                outputdir=os.path.join(robot_path, self.output_dir),
                listener='allure_robotframework;%s' % allure_results_dir,
                reporttitle='Hybrid Robot Report',
                variablefile=variable_files,
                include=[cmd_args['tag']] if cmd_args['tag'] else [],
                suite=cmd_args['suite'],
                test=cmd_args['case']
            )
        elif cmd_args['suite'] and not cmd_args['case']:
            robot.run(
                self.robot_dir,
                consolewidth=80,
                consolecolors='on',
                outputdir=os.path.join(robot_path, self.output_dir),
                listener='allure_robotframework;%s' % allure_results_dir,
                reporttitle='Hybrid Robot Report',
                variablefile=os.listdir(os.path.join(robot_path, self.variables_dir)),
                include=[cmd_args['tag']] if cmd_args['tag'] else [],
                suite=cmd_args['suite']
            )
        elif not cmd_args['suite'] and cmd_args['case']:
            robot.run(
                self.robot_dir,
                consolewidth=80,
                consolecolors='on',
                outputdir=os.path.join(robot_path, self.output_dir),
                listener='allure_robotframework;%s' % allure_results_dir,
                reporttitle='Hybrid Robot Report',
                variablefile=os.listdir(os.path.join(robot_path, self.variables_dir)),
                include=[cmd_args['tag']] if cmd_args['tag'] else [],
                test=cmd_args['case']
            )
        elif not cmd_args['suite'] and not cmd_args['case']:
            robot.run(
                self.robot_dir,
                consolewidth=80,
                consolecolors='on',
                outputdir=os.path.join(robot_path, self.output_dir),
                listener='allure_robotframework;%s' % allure_results_dir,
                reporttitle='Hybrid Robot Report',
                variablefile=variable_files,
                include=[cmd_args['tag']] if cmd_args['tag'] else []
            )
        if os.path.exists(allure_results_dir):
            with open(os.path.join(allure_results_dir, 'environment.properties'), 'w', encoding='utf-8') as f:
                f.write('\n'.join([
                    ''
                ]))
            with open(os.path.join(allure_results_dir, 'executor.json'), 'w', encoding='utf-8') as f:
                f.write(json.dumps({
                    "name": "Hybrid Robot",
                    "type": "hrobot"
                }))

    def generate_report(self):
        if not os.path.exists(self.project_file):
            print(u'这不是一个 hRobot 项目目录')
            return False
        allure_results_path = os.path.join(
            self.work_dir,
            self.robot_dir,
            self.output_dir,
            'allure-results'
        )
        if not os.path.exists(allure_results_path):
            print(u'尚未发现测试用例执行记录，你可以尝试使用 hrobot run 来执行测试用例')
            return False
        os.system('allure generate %s -o report --clean' % allure_results_path)
        os.system('allure open -p 80 report')

    def debug_project(self):
        pass


if __name__ == '__main__':
    print(u'这是 Hybrid Robot 核心内容')
